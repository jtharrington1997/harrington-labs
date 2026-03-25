"""Reformat knife-edge experiment data into clean, standardized workbooks.

Naming convention:
  File:  KE_{material}_{date}.xlsx
  Sheet: {power_mW}mW
  
Each sheet has:
  - Metadata header (experiment, date, laser params, sample info)
  - One sample block per column group, clean headers
  - Consistent column naming: Position (mm) | Ref Power (mW) | std Ref (µW) | ...
"""
import sys
sys.path.insert(0, "src")

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Styling ──────────────────────────────────────────────────────────────

HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1a3a5c")
META_FONT = Font(name="Arial", bold=True, size=10, color="1a3a5c")
META_FILL = PatternFill("solid", fgColor="e8edf2")
DATA_FONT = Font(name="Arial", size=10)
UNIT_FONT = Font(name="Arial", size=9, italic=True, color="666666")
SAMPLE_FONT = Font(name="Arial", bold=True, size=10, color="8b2332")
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Border(bottom=Side(style="thin", color="dddddd"))
NUM_FMT_4 = "0.0000"
NUM_FMT_6 = "0.000000"
NUM_FMT_2 = "0.00"


def style_header(ws, row, c_start, c_end):
    for c in range(c_start, c_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CTR


def style_units(ws, row, c_start, c_end):
    for c in range(c_start, c_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = UNIT_FONT
        cell.alignment = CTR


def write_meta(ws, start_row, fields):
    for i, (label, value) in enumerate(fields):
        r = start_row + i
        cl = ws.cell(row=r, column=1, value=label)
        cl.font = META_FONT
        cl.fill = META_FILL
        cl.alignment = LEFT
        cv = ws.cell(row=r, column=2, value=value)
        cv.font = DATA_FONT
        cv.alignment = LEFT
    return start_row + len(fields) + 1


def write_sample_block(ws, start_row, col_offset, sample_name, headers, units,
                       data_rows, num_fmts=None):
    """Write a single sample block starting at (start_row, col_offset).
    
    Returns (next_free_row, block_width).
    """
    n_cols = len(headers)
    
    # Sample label
    cell = ws.cell(row=start_row, column=col_offset, value=sample_name)
    cell.font = SAMPLE_FONT
    
    # Headers
    hr = start_row + 1
    for j, h in enumerate(headers):
        ws.cell(row=hr, column=col_offset + j, value=h)
    style_header(ws, hr, col_offset, col_offset + n_cols - 1)
    
    # Units
    ur = hr + 1
    for j, u in enumerate(units):
        ws.cell(row=ur, column=col_offset + j, value=u)
    style_units(ws, ur, col_offset, col_offset + n_cols - 1)
    
    # Data
    dr = ur + 1
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=dr + i, column=col_offset + j)
            if val is not None and not (isinstance(val, float) and np.isnan(val)):
                cell.value = val
                if num_fmts and j < len(num_fmts) and num_fmts[j]:
                    cell.number_format = num_fmts[j]
            cell.font = DATA_FONT
            cell.alignment = CTR
            cell.border = THIN
    
    return dr + len(data_rows) + 1, n_cols


def auto_width(ws, max_col=30):
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        widths = []
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    widths.append(len(str(cell.value)))
        if widths:
            ws.column_dimensions[letter].width = min(max(max(widths) + 2, 10), 22)


# ── Data extraction ──────────────────────────────────────────────────────

from harrington_labs.comparison.parsers import detect_and_parse


def group_datasets(filepath):
    """Parse and group datasets by sheet (power level) and sample."""
    datasets = detect_and_parse(filepath)
    groups = {}
    for ds in datasets:
        sheet = ds.metadata.get("sheet", "unknown")
        label = ds.metadata.get("block_label", ds.name.split("—")[0].strip())
        dtype = ds.metadata.get("type", "unknown")
        key = (sheet, label)
        if key not in groups:
            groups[key] = {}
        groups[key][dtype] = ds
    return groups


# ── Si workbook builder ──────────────────────────────────────────────────

def build_si_workbook(filepath_3_19):
    """Build clean Si workbook from 3/19 data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    df_raw = {}
    xls = pd.ExcelFile(filepath_3_19)
    for sname in xls.sheet_names:
        df_raw[sname] = pd.read_excel(xls, sname, header=None)
    
    groups = group_datasets(filepath_3_19)
    
    # Ge info sheet (reference)
    ws_info = wb.create_sheet("Sample Info")
    write_meta(ws_info, 1, [
        ("Experiment", "Knife-Edge Z-Scan — Silicon Samples"),
        ("Date", "2026-03-19"),
        ("Laser", "CW, λ = 2100 nm (assumed)"),
        ("Substrate", "p-type Boron Si, 100 mm, 1000 µm"),
    ])
    r = 7
    ws_info.cell(row=r, column=1, value="Si Sample Inventory").font = SAMPLE_FONT
    r += 1
    samples_si = [
        ("Si-N 9500 Ω·cm", "n-type", "9500 Ω·cm", "~1000 µm"),
        ("Si-P 3–9 Ω", "p-type, Boron", "3–9 Ω", "~1000 µm"),
        ("Si-P 10–20 Ω", "p-type", "10–20 Ω", "~1000 µm"),
        ("Si-P 0.001 Ω", "p-type", "0.001 Ω", "~1000 µm"),
        ("Si Intrinsic", "intrinsic", ">10 kΩ·cm", "~1000 µm"),
    ]
    for label, dtype, res, thick in samples_si:
        ws_info.cell(row=r, column=1, value=label).font = DATA_FONT
        ws_info.cell(row=r, column=2, value=dtype).font = DATA_FONT
        ws_info.cell(row=r, column=3, value=res).font = DATA_FONT
        ws_info.cell(row=r, column=4, value=thick).font = DATA_FONT
        r += 1
    auto_width(ws_info, 4)
    
    # Power level sheets
    power_sheets = [
        ("80 mW", "80mW"),
        ("55 mW", "55mW"),
        ("40 mW", "40mW"),
    ]
    
    # Canonical sample order and clean names
    sample_map = {
        "Si - N 9500": "Si-N 9500Ω",
        "Si-P 3 - 9 ohms": "Si-P 3–9Ω",
        "Si-P 10- 20 ohms": "Si-P 10–20Ω",
        "Si-P  0.001 ohms": "Si-P 0.001Ω",
        "Si Int": "Si Intrinsic",
    }
    sample_order = list(sample_map.keys())
    
    for orig_sheet, clean_sheet in power_sheets:
        ws = wb.create_sheet(clean_sheet)
        power_val = orig_sheet.split()[0]
        
        meta_row = write_meta(ws, 1, [
            ("Experiment", "Knife-Edge Z-Scan — Silicon"),
            ("Date", "2026-03-19"),
            ("Laser Power", f"{power_val} mW"),
            ("Wavelength", "2100 nm"),
            ("Substrate", "p-type Boron Si, 100 mm wafer"),
        ])
        
        col = 1
        for orig_name in sample_order:
            clean_name = sample_map[orig_name]
            key = (orig_sheet, orig_name)
            
            if key not in groups:
                continue
            
            g = groups[key]
            
            # Build data rows from whatever we have
            headers = ["Position"]
            units_row = ["mm"]
            fmts = ["0.00"]
            
            # Get position array from any available dataset
            pos = None
            for dtype in ["norm_trans", "raw_trans", "norm_ref"]:
                if dtype in g:
                    pos = g[dtype].x
                    break
            if pos is None:
                continue
            
            data_cols = [pos.tolist()]
            
            if "norm_ref" in g:
                headers.append("Norm Ref")
                units_row.append("")
                fmts.append(NUM_FMT_6)
                data_cols.append(g["norm_ref"].y.tolist())
            
            if "raw_trans" in g:
                headers.append("Trans Power")
                units_row.append("mW")
                fmts.append(NUM_FMT_2)
                data_cols.append(g["raw_trans"].y.tolist())
            
            if "norm_trans" in g:
                headers.append("Norm Trans")
                units_row.append("")
                fmts.append(NUM_FMT_6)
                data_cols.append(g["norm_trans"].y.tolist())
            
            # Transpose to rows
            n_pts = len(pos)
            rows = []
            for i in range(n_pts):
                row = []
                for dc in data_cols:
                    row.append(dc[i] if i < len(dc) else None)
                rows.append(row)
            
            _, bw = write_sample_block(
                ws, meta_row, col, clean_name,
                headers, units_row, rows, fmts,
            )
            col += bw + 1  # gap column
        
        auto_width(ws, col + 5)
    
    # Also include the full 32-point scan from Sheet2
    ws_full = wb.create_sheet("80mW_FullScan")
    write_meta(ws_full, 1, [
        ("Experiment", "Knife-Edge Z-Scan — Si Full Scan (32 points)"),
        ("Date", "2026-03-19"),
        ("Laser Power", "80 mW (approx)"),
        ("Sample", "p-type Boron Si, 100 mm, 1000 µm"),
    ])
    
    key_full = ("Sheet2", "No sample Power")
    if key_full in groups and "raw_trans" in groups[key_full]:
        ds = groups[key_full]["raw_trans"]
        headers = ["Position", "Ref Power", "Trans Power"]
        units_row = ["mm", "mW", "mW"]
        fmts = ["0.00", NUM_FMT_2, NUM_FMT_2]
        
        # Re-read Sheet2 to get ref power column too
        df2 = df_raw.get("Sheet2", pd.DataFrame())
        ref_data = []
        trans_data = []
        pos_data = []
        for i in range(6, len(df2)):
            row = df2.iloc[i]
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[3]):
                try:
                    pos_data.append(float(row.iloc[0]))
                    ref_data.append(float(row.iloc[1]) if pd.notna(row.iloc[1]) else None)
                    trans_data.append(float(row.iloc[3]))
                except (ValueError, TypeError):
                    pass
        
        rows = list(zip(pos_data, ref_data, trans_data))
        write_sample_block(ws_full, 7, 1, "Si p-type Boron", headers, units_row, rows, fmts)
        auto_width(ws_full, 5)
    
    return wb


# ── Ge workbook builder ──────────────────────────────────────────────────

def build_ge_workbook(filepath_3_19, filepath_3_20):
    """Build clean Ge workbook combining data from both dates."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sample info sheet
    ws_info = wb.create_sheet("Sample Info")
    write_meta(ws_info, 1, [
        ("Experiment", "Knife-Edge Z-Scan — Germanium Samples"),
        ("Dates", "2026-03-19, 2026-03-20"),
        ("Laser", "CW, λ = 2100 nm (assumed)"),
    ])
    r = 6
    ws_info.cell(row=r, column=1, value="Ge Sample Inventory").font = SAMPLE_FONT
    r += 1
    samples_ge = [
        ("Ge Intrinsic", "intrinsic", ">50 Ω·cm", "~500 µm", "specular/mirror"),
        ("Ge-P 5.1 Ω·cm", "p-type", "5.1 Ω·cm", "~700 µm", "specular/specular"),
        ("Ge-P 0.02–0.04 Ω", "p-type", "0.02–0.04 Ω·cm", "317±25 µm", "mirror/pitted"),
        ("Ge-N 42 Ω·cm", "n-type", "42 Ω·cm", "~660 µm", "specular/specular"),
        ("Ge-N 11 Ω·cm", "n-type", "~11 Ω·cm", "~500 µm", "specular/specular"),
    ]
    for row_hdr in ["Name", "Doping", "Resistivity", "Thickness", "Surface"]:
        ws_info.cell(row=r - 1, column=1 + ["Name", "Doping", "Resistivity", "Thickness", "Surface"].index(row_hdr),
                     value=row_hdr).font = META_FONT
    for label, doping, res, thick, surf in samples_ge:
        ws_info.cell(row=r, column=1, value=label).font = DATA_FONT
        ws_info.cell(row=r, column=2, value=doping).font = DATA_FONT
        ws_info.cell(row=r, column=3, value=res).font = DATA_FONT
        ws_info.cell(row=r, column=4, value=thick).font = DATA_FONT
        ws_info.cell(row=r, column=5, value=surf).font = DATA_FONT
        r += 1
    auto_width(ws_info, 5)
    
    # 80 mW data from 3/20 (most complete — has norm trans)
    groups_3_20 = group_datasets(filepath_3_20)
    
    ge_sample_map = {
        "Ge-P 5 ohmn cm": "Ge-P 5.1Ω",
        "Ge-P 0.02-0.04 ohms": "Ge-P 0.02Ω",
        "Ge-N 46 ohm cm": "Ge-N 42Ω",
        "Ge-N 11 ohms": "Ge-N 11Ω",
        "Ge-Intrinsic": "Ge Intrinsic",
    }
    
    ws = wb.create_sheet("80mW_Mar20")
    meta_row = write_meta(ws, 1, [
        ("Experiment", "Knife-Edge Z-Scan — Germanium"),
        ("Date", "2026-03-20"),
        ("Laser Power", "80 mW"),
        ("Wavelength", "2100 nm"),
        ("Notes", "Blank reference + 5 Ge samples"),
    ])
    
    # Blank reference first
    key_blank = ("80 mW", "Ge")
    if key_blank in groups_3_20 and "raw_trans" in groups_3_20[key_blank]:
        ds = groups_3_20[key_blank]["raw_trans"]
        rows = [[p, t] for p, t in zip(ds.x.tolist(), ds.y.tolist())]
        write_sample_block(ws, meta_row, 1, "Blank (no sample)",
                          ["Position", "Trans Power"], ["mm", "mW"],
                          rows, ["0.00", NUM_FMT_2])
    
    col = 5  # start samples after blank
    for orig_name, clean_name in ge_sample_map.items():
        key = ("80 mW", orig_name)
        if key not in groups_3_20:
            continue
        g = groups_3_20[key]
        
        pos = None
        for dtype in ["norm_trans", "raw_trans", "norm_ref"]:
            if dtype in g:
                pos = g[dtype].x
                break
        if pos is None:
            continue
        
        headers = ["Position"]
        units_row = ["mm"]
        fmts = ["0.00"]
        data_cols = [pos.tolist()]
        
        if "norm_ref" in g:
            headers.append("Norm Ref")
            units_row.append("")
            fmts.append(NUM_FMT_6)
            data_cols.append(g["norm_ref"].y.tolist())
        
        if "raw_trans" in g:
            headers.append("Trans Power")
            units_row.append("mW")
            fmts.append(NUM_FMT_4)
            data_cols.append(g["raw_trans"].y.tolist())
        
        if "norm_trans" in g:
            headers.append("Norm Trans")
            units_row.append("")
            fmts.append(NUM_FMT_6)
            data_cols.append(g["norm_trans"].y.tolist())
        
        n_pts = len(pos)
        rows = []
        for i in range(n_pts):
            row = [dc[i] if i < len(dc) else None for dc in data_cols]
            rows.append(row)
        
        _, bw = write_sample_block(ws, meta_row, col, clean_name,
                                   headers, units_row, rows, fmts)
        col += bw + 1
    
    auto_width(ws, col + 5)
    
    # 80 mW Ge from 3/19 (raw power, no normalization)
    groups_3_19 = group_datasets(filepath_3_19)
    
    ge_3_19_map = {
        "x/z, 5 ohmn cm": "Ge-P 5.1Ω",
        "46 ohms": "Ge-N 42Ω",
        "Ge  intrinsic ohms": "Ge Intrinsic",
    }
    
    ws2 = wb.create_sheet("80mW_Mar19")
    meta_row2 = write_meta(ws2, 1, [
        ("Experiment", "Knife-Edge Z-Scan — Germanium"),
        ("Date", "2026-03-19"),
        ("Laser Power", "80 mW"),
        ("Wavelength", "2100 nm"),
        ("Notes", "Raw power data (µW scale for some samples)"),
    ])
    
    col = 1
    for orig_name, clean_name in ge_3_19_map.items():
        key = ("80 mW Ge", orig_name)
        if key not in groups_3_19:
            continue
        g = groups_3_19[key]
        if "raw_trans" not in g:
            continue
        ds = g["raw_trans"]
        
        rows = [[p, t] for p, t in zip(ds.x.tolist(), ds.y.tolist())]
        _, bw = write_sample_block(ws2, meta_row2, col, clean_name,
                                   ["Position", "Trans Power"], ["mm", "µW"],
                                   rows, ["0.00", NUM_FMT_2])
        col += bw + 1
    
    auto_width(ws2, col + 5)
    
    return wb


# ── Main ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    f_3_19 = "/mnt/user-data/uploads/Knife_edge_for_silicon_and_Ge_3_19_26.xlsx"
    f_3_20 = "/mnt/user-data/uploads/Knife_edge_for_silicon_and_Ge_3_20_26.xlsx"
    
    out_dir = Path("/mnt/user-data/outputs")
    
    si_wb = build_si_workbook(f_3_19)
    si_path = out_dir / "KE_Si_2026-03-19.xlsx"
    si_wb.save(si_path)
    print(f"Si workbook: {si_path}")
    for s in si_wb.sheetnames:
        print(f"  Sheet: {s}")
    
    ge_wb = build_ge_workbook(f_3_19, f_3_20)
    ge_path = out_dir / "KE_Ge_2026-03-19_20.xlsx"
    ge_wb.save(ge_path)
    print(f"\nGe workbook: {ge_path}")
    for s in ge_wb.sheetnames:
        print(f"  Sheet: {s}")
