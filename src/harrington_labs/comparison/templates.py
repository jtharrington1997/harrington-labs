"""Generate downloadable experiment data templates.

Creates xlsx workbooks with pre-formatted sheets that the comparison
parsers can ingest directly. Each lab gets a sheet layout matching
its typical measurement workflow. No Streamlit imports.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling ─────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1a3a5c")
_META_FONT = Font(name="Arial", bold=True, size=10, color="1a3a5c")
_META_FILL = PatternFill("solid", fgColor="e8edf2")
_DATA_FONT = Font(name="Arial", size=10)
_UNIT_FONT = Font(name="Arial", size=9, italic=True, color="666666")
_EXAMPLE_FONT = Font(name="Arial", size=10, color="999999")
_NOTE_FONT = Font(name="Arial", size=9, italic=True, color="8b2332")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="cccccc"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _style_unit_row(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _UNIT_FONT
        cell.alignment = _CENTER


def _style_data_rows(ws, row_start: int, row_end: int, col_start: int, col_end: int, example: bool = False):
    font = _EXAMPLE_FONT if example else _DATA_FONT
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER


def _add_metadata_block(ws, start_row: int, fields: list[tuple[str, str]]) -> int:
    """Write a metadata block (label/value pairs) and return next free row."""
    for i, (label, value) in enumerate(fields):
        r = start_row + i
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_l.font = _META_FONT
        cell_l.fill = _META_FILL
        cell_l.alignment = _LEFT
        cell_v = ws.cell(row=r, column=2, value=value)
        cell_v.font = _DATA_FONT
        cell_v.alignment = _LEFT
    return start_row + len(fields) + 1  # one blank row after


def _add_notes(ws, row: int, notes: list[str]) -> int:
    for i, note in enumerate(notes):
        cell = ws.cell(row=row + i, column=1, value=note)
        cell.font = _NOTE_FONT
    return row + len(notes) + 1


def _auto_width(ws, col_start: int = 1, col_end: int = 10):
    for c in range(col_start, col_end + 1):
        letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 28)


# ── Sheet builders ──────────────────────────────────────────────────────────


def _build_zscan_sheet(wb: Workbook, name: str = "Z-Scan"):
    """Open-aperture z-scan / knife-edge measurement template."""
    ws = wb.create_sheet(name)

    # Metadata
    meta = [
        ("Experiment", "Open-Aperture Z-Scan"),
        ("Date", ""),
        ("Operator", ""),
        ("Laser Wavelength (nm)", ""),
        ("Average Power (mW)", ""),
        ("Pulse Width (fs)", ""),
        ("Rep Rate (kHz)", ""),
        ("Beam Diameter (mm)", ""),
        ("Focal Length (mm)", ""),
        ("Sample Material", ""),
        ("Sample Thickness (mm)", ""),
        ("Dopant / Resistivity", ""),
    ]
    data_row = _add_metadata_block(ws, 1, meta)

    # Instructions
    data_row = _add_notes(ws, data_row, [
        "Instructions: Fill in measured data below. Columns match the auto-parser.",
        "The parser recognizes: x/z, Ref Power, std (uW), Transmitted Power, std (uW), Norm Trans.",
        "Add additional sample blocks side-by-side to the right (cols H+).",
    ])

    # Blank reference header
    headers_blank = ["x/z", "Ref Power", "std (uW)", "Transmitted Power", "std (uW)", "Norm Trans"]
    units = ["mm", "mW", "µW", "mW", "µW", ""]
    hr = data_row
    ws.cell(row=hr - 1, column=1, value="Blank (no sample)").font = _META_FONT

    for j, h in enumerate(headers_blank):
        ws.cell(row=hr, column=j + 1, value=h)
    _style_header_row(ws, hr, 1, len(headers_blank))

    for j, u in enumerate(units):
        ws.cell(row=hr + 1, column=j + 1, value=u)
    _style_unit_row(ws, hr + 1, 1, len(units))

    # Example data rows
    example_positions = [28, 26, 24, 20.5, 19, 17.59, 14, 12]
    for i, pos in enumerate(example_positions):
        r = hr + 2 + i
        ws.cell(row=r, column=1, value=pos)
    _style_data_rows(ws, hr + 2, hr + 2 + len(example_positions) - 1, 1, len(headers_blank), example=True)

    # Sample block
    sample_row = hr + 2 + len(example_positions) + 2
    ws.cell(row=sample_row - 1, column=1, value="Sample 1").font = _META_FONT

    headers_sample = ["x/z", "Ref Power", "std (uW)", "Transmitted Power", "std (uW)", "Norm Ref", "Norm Trans"]
    for j, h in enumerate(headers_sample):
        ws.cell(row=sample_row, column=j + 1, value=h)
    _style_header_row(ws, sample_row, 1, len(headers_sample))

    units_sample = ["mm", "mW", "µW", "mW", "µW", "", ""]
    for j, u in enumerate(units_sample):
        ws.cell(row=sample_row + 1, column=j + 1, value=u)
    _style_unit_row(ws, sample_row + 1, 1, len(units_sample))

    for i, pos in enumerate(example_positions):
        ws.cell(row=sample_row + 2 + i, column=1, value=pos)
    _style_data_rows(ws, sample_row + 2, sample_row + 2 + len(example_positions) - 1, 1, len(headers_sample), example=True)

    _auto_width(ws, 1, len(headers_sample))
    return ws


def _build_li_curve_sheet(wb: Workbook, name: str = "L-I Curve"):
    """Diode L-I characteristic measurement template."""
    ws = wb.create_sheet(name)

    meta = [
        ("Experiment", "L-I Curve Measurement"),
        ("Date", ""),
        ("Operator", ""),
        ("Diode Model", ""),
        ("Wavelength (nm)", ""),
        ("Heatsink Temperature (°C)", ""),
        ("Threshold Current (A)", ""),
    ]
    data_row = _add_metadata_block(ws, 1, meta)

    data_row = _add_notes(ws, data_row, [
        "Instructions: Record current, optical power, and voltage at each operating point.",
    ])

    headers = ["Current (A)", "Optical Power (W)", "std (W)", "Voltage (V)", "Junction Temp (°C)"]
    for j, h in enumerate(headers):
        ws.cell(row=data_row, column=j + 1, value=h)
    _style_header_row(ws, data_row, 1, len(headers))
    _style_data_rows(ws, data_row + 1, data_row + 20, 1, len(headers), example=True)

    _auto_width(ws, 1, len(headers))
    return ws


def _build_fiber_sheet(wb: Workbook, name: str = "Fiber Laser"):
    """Fiber laser power evolution measurement template."""
    ws = wb.create_sheet(name)

    meta = [
        ("Experiment", "Fiber Laser Power Evolution"),
        ("Date", ""),
        ("Operator", ""),
        ("Fiber Type", ""),
        ("Core Diameter (µm)", ""),
        ("Pump Wavelength (nm)", ""),
        ("Pump Power (W)", ""),
        ("Signal Wavelength (nm)", ""),
    ]
    data_row = _add_metadata_block(ws, 1, meta)

    data_row = _add_notes(ws, data_row, [
        "Instructions: Record signal power at tapped points along the fiber.",
    ])

    headers = ["Position (m)", "Signal Power (W)", "std (W)", "Pump Remaining (W)", "std (W)"]
    for j, h in enumerate(headers):
        ws.cell(row=data_row, column=j + 1, value=h)
    _style_header_row(ws, data_row, 1, len(headers))
    _style_data_rows(ws, data_row + 1, data_row + 15, 1, len(headers), example=True)

    _auto_width(ws, 1, len(headers))
    return ws


def _build_beam_profile_sheet(wb: Workbook, name: str = "Beam Profile"):
    """Beam radius / irradiance propagation measurement template."""
    ws = wb.create_sheet(name)

    meta = [
        ("Experiment", "Beam Propagation / Profile"),
        ("Date", ""),
        ("Operator", ""),
        ("Wavelength (nm)", ""),
        ("Power (W)", ""),
        ("Beam Diameter at Source (mm)", ""),
        ("Measurement Method", ""),  # knife-edge, camera, etc.
    ]
    data_row = _add_metadata_block(ws, 1, meta)

    data_row = _add_notes(ws, data_row, [
        "Instructions: Record beam radius (or diameter) and optional irradiance at each distance.",
    ])

    headers = ["Distance (m)", "Beam Radius (cm)", "std (cm)", "Irradiance (W/cm²)", "std (W/cm²)"]
    for j, h in enumerate(headers):
        ws.cell(row=data_row, column=j + 1, value=h)
    _style_header_row(ws, data_row, 1, len(headers))
    _style_data_rows(ws, data_row + 1, data_row + 20, 1, len(headers), example=True)

    _auto_width(ws, 1, len(headers))
    return ws


def _build_spectrum_sheet(wb: Workbook, name: str = "Spectrum"):
    """Spectral measurement template (PL, absorption, reflectance)."""
    ws = wb.create_sheet(name)

    meta = [
        ("Experiment", "Spectral Measurement"),
        ("Date", ""),
        ("Operator", ""),
        ("Measurement Type", ""),  # PL, absorption, reflectance, transmittance
        ("Excitation Wavelength (nm)", ""),
        ("Sample", ""),
        ("Temperature (K)", ""),
        ("Spectrometer Model", ""),
        ("Integration Time (ms)", ""),
    ]
    data_row = _add_metadata_block(ws, 1, meta)

    data_row = _add_notes(ws, data_row, [
        "Instructions: Record wavelength and signal intensity. Add uncertainty if available.",
        "For reflectance/transmittance, values should be normalized 0–1.",
    ])

    headers = ["Wavelength (nm)", "Intensity (a.u.)", "std", "Reflectance", "std"]
    for j, h in enumerate(headers):
        ws.cell(row=data_row, column=j + 1, value=h)
    _style_header_row(ws, data_row, 1, len(headers))
    _style_data_rows(ws, data_row + 1, data_row + 50, 1, len(headers), example=True)

    _auto_width(ws, 1, len(headers))
    return ws


def _build_generic_sheet(wb: Workbook, name: str = "Generic"):
    """Generic two-column x/y measurement template."""
    ws = wb.create_sheet(name)

    meta = [
        ("Experiment", ""),
        ("Date", ""),
        ("Operator", ""),
        ("Notes", ""),
    ]
    data_row = _add_metadata_block(ws, 1, meta)

    data_row = _add_notes(ws, data_row, [
        "Instructions: Use this sheet for any x-y measurement not covered by other tabs.",
        "The parser will auto-detect numeric column pairs.",
    ])

    headers = ["x", "y", "y_uncertainty"]
    for j, h in enumerate(headers):
        ws.cell(row=data_row, column=j + 1, value=h)
    _style_header_row(ws, data_row, 1, len(headers))
    _style_data_rows(ws, data_row + 1, data_row + 30, 1, len(headers), example=True)

    _auto_width(ws, 1, len(headers))
    return ws


# ── CSV templates ───────────────────────────────────────────────────────────


_CSV_TEMPLATES = {
    "zscan": "x/z,Ref Power,std (uW),Transmitted Power,std (uW),Norm Trans\nmm,mW,uW,mW,uW,\n",
    "li_curve": "Current (A),Optical Power (W),std (W),Voltage (V),Junction Temp (C)\n",
    "fiber": "Position (m),Signal Power (W),std (W),Pump Remaining (W),std (W)\n",
    "beam_profile": "Distance (m),Beam Radius (cm),std (cm),Irradiance (W/cm2),std (W/cm2)\n",
    "spectrum": "Wavelength (nm),Intensity (a.u.),std,Reflectance,std\n",
    "generic": "x,y,y_uncertainty\n",
}


# ── Public API ──────────────────────────────────────────────────────────────


def generate_full_template(output_path: str | Path = "data/templates/experiment_template.xlsx") -> Path:
    """Generate an xlsx workbook with all lab template sheets."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    _build_zscan_sheet(wb, "Z-Scan")
    _build_li_curve_sheet(wb, "L-I Curve")
    _build_fiber_sheet(wb, "Fiber Laser")
    _build_beam_profile_sheet(wb, "Beam Profile")
    _build_spectrum_sheet(wb, "Spectrum")
    _build_generic_sheet(wb, "Generic")

    wb.save(output_path)
    return output_path


def generate_lab_template(
    lab_type: str,
    output_path: str | Path | None = None,
) -> Path:
    """Generate a single-sheet template for a specific lab.

    lab_type: one of 'zscan', 'li_curve', 'fiber', 'beam_profile', 'spectrum', 'generic'
    """
    builders = {
        "zscan": _build_zscan_sheet,
        "li_curve": _build_li_curve_sheet,
        "fiber": _build_fiber_sheet,
        "beam_profile": _build_beam_profile_sheet,
        "spectrum": _build_spectrum_sheet,
        "generic": _build_generic_sheet,
    }
    if lab_type not in builders:
        raise ValueError(f"Unknown lab_type: {lab_type}. Options: {list(builders)}")

    if output_path is None:
        output_path = Path(f"data/templates/{lab_type}_template.xlsx")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    builders[lab_type](wb)
    wb.save(output_path)
    return output_path


def get_csv_template(lab_type: str) -> str:
    """Return a CSV header string template for a given lab type."""
    if lab_type not in _CSV_TEMPLATES:
        raise ValueError(f"Unknown lab_type: {lab_type}. Options: {list(_CSV_TEMPLATES)}")
    return _CSV_TEMPLATES[lab_type]
